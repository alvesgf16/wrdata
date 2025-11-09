"""
Excel writer module for champion data output.

This module provides functionality for writing champion data to Excel
workbooks. It implements the Writer abstract base class to handle the
specific requirements of Excel file output, including worksheet
creation and table formatting.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ...domain.models.analyzed_champion import AnalyzedChampion
from ...exceptions import OutputError
from .champion_serializer import ChampionSerializer
from .writer import Writer


class XlsxWriter(Writer):
    """A class for writing champion data to Excel workbooks.

    This class implements the Writer abstract base class to write
    champion data to Excel format. It creates a single workbook with
    multiple worksheets, one for each tier, and formats the data as
    tables with proper headers.

    The output file is created in the 'res' directory with the
    format: 'output_filename.xlsx'

    Methods:
    ---
    write():
        Write champion data to an Excel workbook with multiple worksheets.
    """

    def write(self, data: list[AnalyzedChampion]) -> None:
        """Write champion data to an Excel workbook with worksheets.

        This method creates a single Excel workbook and writes the
        champion data to separate worksheets for each tier. Champions
        are automatically grouped by their tier attribute.

        Args:
            data (list[AnalyzedChampion]): A flat list of
                AnalyzedChampion objects.

        Raises:
            OutputError: If there are issues creating or writing to the
                Excel file
        """
        try:
            output_file = self._create_output_file().with_suffix(".xlsx")
            workbook = Workbook()

            # Remove default sheet only if we have data to write
            if data and "Sheet" in workbook.sheetnames:
                default_sheet = workbook["Sheet"]
                workbook.remove(default_sheet)

            # Group champions by tier
            if data:
                tier_groups: dict[str, list[AnalyzedChampion]] = {}
                for champion in data:
                    tier_name = (
                        champion.tier.name if champion.tier else "Unranked"
                    )
                    if tier_name not in tier_groups:
                        tier_groups[tier_name] = []
                    tier_groups[tier_name].append(champion)

                # Write a worksheet for each tier group
                for tier_name in self._tiers:
                    if tier_name in tier_groups:
                        self.__write_tier_worksheet(
                            workbook, tier_name, tier_groups[tier_name]
                        )

            workbook.save(output_file)
        except Exception as e:
            raise OutputError(
                "Failed to write champion data to Excel file", details=str(e)
            ) from e

    def __write_tier_worksheet(
        self,
        workbook: Workbook,
        tier_name: str,
        tier_data: list[AnalyzedChampion],
    ) -> None:
        """Create and populate a worksheet for a single tier.

        This method creates a new worksheet in the workbook for a
        specific tier and formats the champion data as a table with
        proper headers. The table includes all champion statistics
        defined in the headers.

        Args:
            workbook (Workbook): The Excel workbook to add the
                worksheet to.
            tier_name (str): The name of the tier being written.
            tier_data (list[AnalyzedChampion]): List of AnalyzedChampion
                objects belonging to this tier.
        """
        worksheet = workbook.create_sheet(title=tier_name)

        # Write headers
        for col_idx, header in enumerate(self._headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, champion in enumerate(tier_data, start=2):
            row_data = ChampionSerializer.to_csv_row(champion)
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Create table if there's data
        if tier_data:
            last_row = len(tier_data) + 1
            last_col = get_column_letter(len(self._headers))

            table = Table(
                displayName=f"Table_{tier_name}",
                ref=f"A1:{last_col}{last_row}",
            )

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style

            worksheet.add_table(table)
