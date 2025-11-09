"""
Excel-based champion data reader implementation.

This module provides a concrete implementation of the ChampionReader port
for reading champion data from Excel workbooks.
"""

from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ...data import Champion, Lane
from .reader import ChampionReader


class ExcelChampionReader(ChampionReader):
    """Excel-based implementation of the ChampionReader port.

    Reads champion data from Excel workbooks and reconstructs Champion
    objects organized by tier. This reader is designed to work with Excel
    files created by ExcelChampionRepository.

    Requires openpyxl to be installed. If not installed, will raise
    ImportError when attempting to read.
    """

    def __init__(self, filepath: str = "champions.xlsx") -> None:
        """Initialize the Excel champion reader.

        Args:
            filepath: The path to the Excel file to read from.
        """
        self._filepath = Path(filepath)

    def read(self) -> list[list[Champion]]:
        """Read champion data from the Excel file.

        Parses each worksheet as a tier and reconstructs Champion objects
        organized by tier. Each worksheet represents one tier of champions.

        Returns:
            A list of lists containing Champion objects, where each
            inner list represents champions from one worksheet (tier).

        Raises:
            ImportError: If openpyxl is not installed.
            FileNotFoundError: If the Excel file doesn't exist.
            ValueError: If the Excel data is malformed.
            Exception: For other reading errors.
        """
        if not self._filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self._filepath}")

        try:
            workbook = load_workbook(self._filepath, data_only=True)
            champions_by_tier: list[list[Champion]] = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                tier_champions = self._read_worksheet(sheet)
                if tier_champions:
                    champions_by_tier.append(tier_champions)

            workbook.close()
            return champions_by_tier

        except Exception as e:
            raise Exception(
                f"Failed to read champions from Excel file "
                f"{self._filepath}: {e}"
            ) from e

    def _read_worksheet(self, sheet: Worksheet) -> list[Champion]:
        """Read champions from a single worksheet.

        Args:
            sheet: An openpyxl worksheet object.

        Returns:
            A list of Champion objects from this worksheet.
        """
        champions: list[Champion] = []

        # Get header row to map column indices
        headers = self.get_headers(sheet)

        FIRST_DATA_ROW = 2

        for row in sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
            if row[0] is None:
                continue

            try:
                champion = self._deserialize_champion(headers, row)
                champions.append(champion)
            except ValueError as e:
                print(f"Warning: Skipping invalid row: {row}. Error: {e}")
                continue

        return champions

    def get_headers(self, sheet: Worksheet) -> list[Any]:
        return [cell.value for cell in sheet[1]]

    def _deserialize_champion(
        self, headers: Sequence[str], row: Sequence[Any]
    ) -> Champion:
        """Deserialize an Excel row into a Champion object.

        Args:
            headers: List of column headers from the worksheet.
            row: A tuple representing one Excel row.

        Returns:
            A Champion object reconstructed from the Excel data.

        Raises:
            ValueError: If the row data is invalid.
        """
        try:
            data = dict(zip(headers, row))

            return Champion(
                name=str(data["Champion"]),
                lane=Lane(str(data["Lane"])),
                win_rate=float(data["Win Rate"]),
                pick_rate=float(data["Pick Rate"]),
                ban_rate=float(data["Ban Rate"]),
                ranked_tier=str(data["Ranked Tier"]),
            )
        except (KeyError, ValueError, TypeError) as e:
            raise ValueError(
                f"Invalid Excel row data: {row}. Error: {e}"
            ) from e
