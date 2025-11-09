"""
Excel-based champion repository implementation.

This module provides a concrete implementation of the ChampionRepository port
for persisting champion data to Excel workbooks.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ....exceptions import OutputError
from ...domain import Champion
from ...domain.ports.repositories import ChampionRepository


class ExcelChampionRepository(ChampionRepository):
    """Excel-based implementation of the ChampionRepository port."""

    def __init__(self, filepath: str = "champions.xlsx") -> None:
        """Initialize the Excel champion repository.

        Args:
            filepath: The path where the Excel file will be saved.
        """
        self._filepath = Path(filepath)
        self._headers = [
            "Lane",
            "Champion",
            "Win Rate",
            "Pick Rate",
            "Ban Rate",
        ]
        self._tier_names = ["Tier 1", "Tier 2", "Tier 3", "Tier 4"]

    def save(self, champions: list[list[Champion]]) -> None:
        """Save champion data to an Excel workbook.

        Creates a workbook with separate worksheets for each tier.

        Args:
            champions: A list of lists containing Champion objects,
                where each inner list represents a tier.

        Raises:
            OutputError: If there are issues writing to the Excel file.
        """
        try:
            self._filepath.parent.mkdir(parents=True, exist_ok=True)

            workbook = Workbook()

            # Remove default sheet
            if "Sheet" in workbook.sheetnames:
                default_sheet = workbook["Sheet"]
                workbook.remove(default_sheet)

            for tier_name, tier_data in zip(self._tier_names, champions):
                self._write_tier_worksheet(workbook, tier_name, tier_data)

            workbook.save(self._filepath)

        except Exception as e:
            raise OutputError(
                f"Failed to write champions to Excel file "
                f"{self._filepath}",
                details=str(e),
            ) from e

    def _write_tier_worksheet(
        self,
        workbook: Workbook,
        tier_name: str,
        tier_data: list[Champion],
    ) -> None:
        """Create and populate a worksheet for a single tier.

        Args:
            workbook: The Excel workbook to add the worksheet to.
            tier_name: The name of the tier worksheet.
            tier_data: List of Champion objects for this tier.
        """
        worksheet = workbook.create_sheet(title=tier_name)

        # Write headers
        for col_idx, header in enumerate(self._headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, champion in enumerate(tier_data, start=2):
            row_data = self._serialize_champion(champion)
            for col_idx, value in enumerate(row_data, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

        # Create table if there's data
        if tier_data:
            last_row = len(tier_data) + 1
            last_col = get_column_letter(len(self._headers))

            table = Table(
                displayName=f"Table_{tier_name.replace(' ', '_')}",
                ref=f"A1:{last_col}{last_row}",
            )

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style

            worksheet.add_table(table)

    def _serialize_champion(self, champion: Champion) -> list[str | float]:
        """Serialize a champion to Excel row format.

        Args:
            champion: The Champion object to serialize.

        Returns:
            A list containing the champion's data fields.
        """
        return [
            champion.lane.value,
            champion.name,
            round(champion.win_rate, 4),
            round(champion.pick_rate, 4),
            round(champion.ban_rate, 4),
        ]
