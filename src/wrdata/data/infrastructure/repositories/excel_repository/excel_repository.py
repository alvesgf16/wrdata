"""
Excel-based champion repository implementation.

This module provides a concrete implementation of the ChampionRepository port
for persisting champion data to Excel workbooks.
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .....exceptions import OutputError
from ....domain import Champion
from ..base_repository import BaseChampionRepository


class ExcelChampionRepository(BaseChampionRepository):
    """Excel-based implementation of the ChampionRepository port."""

    def __init__(self, filepath: str = "champions.xlsx") -> None:
        """Initialize the Excel champion repository.

        Args:
            filepath: The path where the Excel file will be saved.
        """
        super().__init__(filepath)

    def save(self, champions: list[Champion]) -> None:
        """Save champion data to an Excel workbook.

        Creates a single worksheet with all champion data.

        Args:
            champions: A flat list of Champion objects.

        Raises:
            OutputError: If there are issues writing to the Excel file.
        """
        try:
            self._ensure_directory_exists()

            self._write(champions)

        except Exception as e:
            raise OutputError(
                f"Failed to write champions to Excel file "
                f"{self._filepath}",
                details=str(e),
            ) from e

    def _write(self, champions: list[Champion]) -> None:
        workbook = Workbook()
        worksheet = self._create_writer(workbook)

        self._write_headers(worksheet)
        self._write_data(champions, worksheet)

        workbook.save(self._filepath)

    def _create_writer(self, workbook: Workbook) -> Worksheet:
        worksheet = workbook.active

        if worksheet is None:
            worksheet = workbook.create_sheet("Champions")
        else:
            worksheet.title = "Champions"

        return worksheet

    def _write_headers(self, worksheet: Worksheet) -> None:
        FIRST_COLUMN_INDEX = 1

        for column_index, header in enumerate(
            self._headers, start=FIRST_COLUMN_INDEX
        ):
            worksheet.cell(
                row=FIRST_COLUMN_INDEX, column=column_index, value=header
            )

    def _write_data(
        self, champions: list[Champion], worksheet: Worksheet
    ) -> None:
        FIRST_DATA_ROW_INDEX = 2
        FIRST_COLUMN_INDEX = 1

        for row_index, champion in enumerate(
            champions, start=FIRST_DATA_ROW_INDEX
        ):
            row_data = self._serialize_champion(champion)
            for column_index, value in enumerate(
                row_data, start=FIRST_COLUMN_INDEX
            ):
                worksheet.cell(row=row_index, column=column_index, value=value)
