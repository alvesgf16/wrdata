"""
Tests for the Excel Champion Repository.
"""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.wrdata.data import Champion, Lane, RankedTier
from src.wrdata.data.infrastructure.repositories.excel_repository import (
    ExcelChampionRepository,
)
from src.wrdata.exceptions import OutputError


@pytest.fixture
def sample_champions() -> list[Champion]:
    """Create sample champions for testing."""
    return [
        Champion(
            name="Darius",
            lane=Lane.TOP,
            win_rate=52.3456,
            pick_rate=8.7654,
            ban_rate=15.2345,
            ranked_tier=RankedTier.DIAMOND_PLUS,
        ),
        Champion(
            name="Lee Sin",
            lane=Lane.JUNGLE,
            win_rate=48.5678,
            pick_rate=12.3456,
            ban_rate=20.1234,
            ranked_tier=RankedTier.SOVEREIGN,
        ),
    ]


@pytest.fixture
def temp_excel_path(tmp_path: Path) -> Path:
    """Create a temporary Excel file path."""
    return tmp_path / "test_champions.xlsx"


def test_excel_repository_initialization() -> None:
    """Test Excel repository initialization with default filepath."""
    repo = ExcelChampionRepository()
    assert repo._filepath == Path("champions.xlsx")


def test_excel_repository_custom_filepath() -> None:
    """Test Excel repository initialization with custom filepath."""
    custom_path = "custom/path/champions.xlsx"
    repo = ExcelChampionRepository(filepath=custom_path)
    assert repo._filepath == Path(custom_path)


def test_save_creates_excel_file(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that save creates an Excel file."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(sample_champions)

    assert temp_excel_path.exists()
    assert temp_excel_path.is_file()


def test_save_creates_champions_worksheet(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that a worksheet named 'Champions' is created."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(sample_champions)

    workbook = load_workbook(temp_excel_path)
    assert "Champions" in workbook.sheetnames
    active_sheet = workbook.active
    assert active_sheet is not None
    assert active_sheet.title == "Champions"


def test_save_writes_headers(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that headers are written correctly."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(sample_champions)

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]

    expected_headers = [
        "Lane",
        "Champion",
        "Win Rate",
        "Pick Rate",
        "Ban Rate",
        "Ranked Tier",
    ]

    headers = [worksheet.cell(row=1, column=col).value for col in range(1, 7)]
    assert headers == expected_headers


def test_save_writes_champion_data(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that champion data is written correctly."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(sample_champions)

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]

    # Check first champion (row 2)
    assert worksheet.cell(row=2, column=1).value == "Top"
    assert worksheet.cell(row=2, column=2).value == "Darius"
    assert worksheet.cell(row=2, column=3).value == 52.3456
    assert worksheet.cell(row=2, column=4).value == 8.7654
    assert worksheet.cell(row=2, column=5).value == 15.2345
    assert worksheet.cell(row=2, column=6).value == "Diamond+"

    # Check second champion (row 3)
    assert worksheet.cell(row=3, column=1).value == "Jungle"
    assert worksheet.cell(row=3, column=2).value == "Lee Sin"
    assert worksheet.cell(row=3, column=3).value == 48.5678
    assert worksheet.cell(row=3, column=4).value == 12.3456
    assert worksheet.cell(row=3, column=5).value == 20.1234
    assert worksheet.cell(row=3, column=6).value == "Sovereign"


def test_save_with_empty_list(temp_excel_path: Path) -> None:
    """Test saving an empty list of champions."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save([])

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]

    # Check headers exist
    assert worksheet.cell(row=1, column=1).value == "Lane"

    # Check no data rows
    assert worksheet.cell(row=2, column=1).value is None


def test_save_creates_nested_directories(tmp_path: Path) -> None:
    """Test that save creates nested directories if they don't exist."""
    nested_path = tmp_path / "nested" / "dir" / "champions.xlsx"
    repo = ExcelChampionRepository(filepath=str(nested_path))

    repo.save(
        [
            Champion(
                name="Test",
                lane=Lane.MID,
                win_rate=50.0,
                pick_rate=5.0,
                ban_rate=2.0,
                ranked_tier=RankedTier.DIAMOND_PLUS,
            )
        ]
    )

    assert nested_path.exists()
    assert nested_path.parent.exists()


def test_save_overwrites_existing_file(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that save overwrites an existing file."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))

    # First save
    repo.save(sample_champions[:1])

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]
    assert worksheet.cell(row=2, column=1).value == "Top"
    assert worksheet.cell(row=3, column=1).value is None

    # Second save should overwrite
    repo.save(sample_champions)

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]
    assert worksheet.cell(row=2, column=1).value == "Top"
    assert worksheet.cell(row=3, column=1).value == "Jungle"


def test_save_raises_output_error_on_invalid_path() -> None:
    """Test that save raises OutputError on invalid path."""
    # Use an invalid path (e.g., path with invalid characters)
    invalid_path = "/\x00invalid/path.xlsx"
    repo = ExcelChampionRepository(filepath=invalid_path)

    with pytest.raises(OutputError) as exc_info:
        repo.save([])

    assert "Failed to write champions to Excel file" in str(exc_info.value)


def test_multiple_champions_same_lane(temp_excel_path: Path) -> None:
    """Test saving multiple champions from the same lane."""
    champions = [
        Champion(
            name="Darius",
            lane=Lane.TOP,
            win_rate=52.0,
            pick_rate=8.0,
            ban_rate=15.0,
            ranked_tier=RankedTier.DIAMOND_PLUS,
        ),
        Champion(
            name="Garen",
            lane=Lane.TOP,
            win_rate=51.0,
            pick_rate=7.0,
            ban_rate=5.0,
            ranked_tier=RankedTier.DIAMOND_PLUS,
        ),
        Champion(
            name="Sett",
            lane=Lane.TOP,
            win_rate=50.0,
            pick_rate=6.0,
            ban_rate=8.0,
            ranked_tier=RankedTier.MASTER_PLUS,
        ),
    ]

    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(champions)

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]

    # All champions should be in one worksheet
    assert worksheet.cell(row=2, column=2).value == "Darius"
    assert worksheet.cell(row=3, column=2).value == "Garen"
    assert worksheet.cell(row=4, column=2).value == "Sett"


def test_unicode_champion_names(temp_excel_path: Path) -> None:
    """Test that unicode champion names are handled correctly."""
    unicode_champion = Champion(
        name="疾风剑豪",  # Yasuo in Chinese
        lane=Lane.MID,
        win_rate=50.5,
        pick_rate=15.2,
        ban_rate=10.3,
        ranked_tier=RankedTier.MASTER_PLUS,
    )

    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save([unicode_champion])

    workbook = load_workbook(temp_excel_path)
    worksheet = workbook["Champions"]

    assert worksheet.cell(row=2, column=2).value == "疾风剑豪"


def test_worksheet_has_no_default_sheet(
    sample_champions: list[Champion], temp_excel_path: Path
) -> None:
    """Test that the default 'Sheet' worksheet is not present."""
    repo = ExcelChampionRepository(filepath=str(temp_excel_path))
    repo.save(sample_champions)

    workbook = load_workbook(temp_excel_path)

    # Only "Champions" sheet should exist
    assert "Champions" in workbook.sheetnames
    assert len(workbook.sheetnames) == 1
